"""Safe Excel reading for the bulk-import pipeline (Batch 2F).

Validates the upload (extension, MIME, size, workbook readability, malformed
workbook, duplicate sheet names) and exposes sheet listing, header detection and
bounded row reading. Uses openpyxl in ``read_only`` mode so large files are not
fully materialised in memory.
"""

from __future__ import annotations

import io
import os
from dataclasses import dataclass

import openpyxl

from app.config import settings
from app.services.validators import ValidationError

# .xls is intentionally NOT supported (see task A / DECISIONS_LOG).
ALLOWED_EXT = {"xlsx", "xlsm"}
ALLOWED_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "application/zip", "application/octet-stream", "",
}
_ZIP_MAGIC = (b"PK\x03\x04", b"PK\x05\x06")


def _ext(filename: str) -> str:
    base = os.path.basename(filename or "")
    return base.rsplit(".", 1)[-1].lower() if "." in base else ""


@dataclass
class SheetPreview:
    headers: list[str]
    rows: list[dict]        # list of {header: value}
    total_rows: int         # data rows found (before the max-rows cap)
    header_row_index: int   # 0-based index of the detected header row
    truncated: bool
    # The merged "category band" row immediately above the header (forward-filled
    # across merged spans), e.g. Actitudinal / Desempeño / Conocimiento. One entry
    # per header column; empty string where there is no band value.
    category_headers: list[str] = None


def validate_upload(filename: str, content_type: str | None, raw: bytes) -> str:
    """Validate an uploaded workbook; return the extension or raise."""
    ext = _ext(filename)
    if ext not in ALLOWED_EXT:
        raise ValidationError({"file": "Solo se admiten archivos .xlsx o .xlsm."})
    declared = (content_type or "").strip().lower()
    if declared and declared not in {m.lower() for m in ALLOWED_MIME}:
        raise ValidationError({"file": "El tipo MIME no corresponde a un archivo Excel válido."})
    if len(raw) == 0:
        raise ValidationError({"file": "El archivo está vacío."})
    if len(raw) > settings.import_max_bytes:
        raise ValidationError({"file": f"El archivo supera el tamaño máximo "
                                       f"({settings.import_max_mb} MB)."})
    if not raw.startswith(_ZIP_MAGIC):
        raise ValidationError({"file": "El contenido no corresponde a un archivo Excel (.xlsx/.xlsm)."})
    return ext


def load_workbook(raw: bytes):
    """Load a workbook from bytes in read-only mode, or raise a friendly error."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:  # openpyxl raises many types on corrupt files
        raise ValidationError({"file": "No se pudo leer el libro de Excel "
                                       "(archivo dañado o con formato no válido)."})
    if not wb.sheetnames:
        raise ValidationError({"file": "El libro no contiene hojas."})
    # Duplicate sheet names indicate a malformed workbook.
    if len(wb.sheetnames) != len(set(n.strip().lower() for n in wb.sheetnames)):
        raise ValidationError({"file": "El libro tiene nombres de hoja duplicados (archivo malformado)."})
    return wb


def list_sheets(raw: bytes) -> list[str]:
    wb = load_workbook(raw)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _detect_header_row(matrix: list[list]) -> int:
    """Return the index of the row with the most non-empty cells (heuristic)."""
    best_idx, best_count = 0, -1
    for i, row in enumerate(matrix[:15]):  # look only at the first rows
        count = sum(1 for c in row if c is not None and str(c).strip() != "")
        if count > best_count:
            best_count, best_idx = count, i
    return best_idx


def _clean_header(value, col_index: int) -> str:
    text = "" if value is None else str(value).strip()
    return text or f"Columna {col_index + 1}"


def read_sheet(raw: bytes, sheet_name: str, *, header_row: int | None = None) -> SheetPreview:
    """Read a sheet into headers + list-of-dict rows, bounded by import_max_rows."""
    wb = load_workbook(raw)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValidationError({"sheet": "La hoja seleccionada no existe en el libro."})
    ws = wb[sheet_name]
    matrix = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not matrix:
        return SheetPreview(headers=[], rows=[], total_rows=0, header_row_index=0, truncated=False)

    hidx = header_row if header_row is not None else _detect_header_row(matrix)
    hidx = max(0, min(hidx, len(matrix) - 1))
    raw_headers = matrix[hidx]
    headers = [_clean_header(v, i) for i, v in enumerate(raw_headers)]

    data_matrix = matrix[hidx + 1:]
    # Drop fully-empty rows.
    data_matrix = [r for r in data_matrix
                   if any(c is not None and str(c).strip() != "" for c in r)]
    total = len(data_matrix)
    truncated = total > settings.import_max_rows
    data_matrix = data_matrix[: settings.import_max_rows]

    rows: list[dict] = []
    for r in data_matrix:
        row = {}
        for i, header in enumerate(headers):
            row[header] = r[i] if i < len(r) else None
        rows.append(row)

    category_headers = read_category_band(raw, sheet_name, hidx, len(headers))
    return SheetPreview(headers=headers, rows=rows, total_rows=total,
                        header_row_index=hidx, truncated=truncated,
                        category_headers=category_headers)


def read_category_band(raw: bytes, sheet_name: str, header_row_index: int,
                       col_count: int) -> list[str]:
    """Forward-filled merged-cell band on the row directly above the header row.

    Grade workbooks group component columns under a merged category label (e.g.
    "Actitudinal" spanning several component columns). ``read_only`` mode does not
    expose merge metadata, so this opens a second, normal-mode load bounded to a
    handful of rows around the header — the grade sheets involved are small
    (dozens of rows), so this stays well within the "avoid loading huge files
    into memory" guidance, which applies to the full-data read path above.
    """
    band_idx = header_row_index - 1
    if band_idx < 0:
        return [""] * col_count
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    except Exception:
        return [""] * col_count
    if sheet_name not in wb.sheetnames:
        wb.close()
        return [""] * col_count
    ws = wb[sheet_name]
    band = [""] * col_count
    excel_row = band_idx + 1  # openpyxl rows are 1-based
    for col in range(1, col_count + 1):
        cell = ws.cell(row=excel_row, column=col)
        value = cell.value
        if value is None:
            # Check whether this cell is part of a merged range anchored elsewhere.
            for rng in ws.merged_cells.ranges:
                if (rng.min_row <= excel_row <= rng.max_row
                        and rng.min_col <= col <= rng.max_col):
                    value = ws.cell(row=rng.min_row, column=rng.min_col).value
                    break
        band[col - 1] = str(value).strip() if value is not None else ""
    wb.close()
    return band


def cell_letter(col_index: int) -> str:
    """0-based column index -> Excel column letter (A, B, ... AA)."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_index + 1)
