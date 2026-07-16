"""Excel (openpyxl) and PDF (fpdf2) generation — local, offline, no paid services.

Two primitives power every report export:

* ``excel_from_table`` — a styled ``.xlsx`` workbook with a header block
  (title, generation date, filters, generator) and a data table.
* ``pdf_from_table`` — a printable ``.pdf`` with the same header block and an
  auto-wrapped table.

``document_to_pdf`` renders a single formal document in the institutional letter
layout modelled on the attached resignation reference.
"""

from __future__ import annotations

import io
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings

_HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=14, bold=True, color="1F3B57")

# Characters outside latin-1 that the PDF core fonts cannot encode.
_REPLACEMENTS = {
    "—": "-", "–": "-", "‘": "'", "’": "'",
    "“": '"', "”": '"', "…": "...", " ": " ",
    "•": "-", "→": "->",
}


def _mc(pdf, h: float, text: str, align: str = "L") -> None:
    """multi_cell that always starts at the left margin and advances one line.

    fpdf2's ``multi_cell`` leaves x at the right margin by default, which makes a
    subsequent full-width ``multi_cell`` compute a zero width and raise. Resetting
    x first and forcing ``new_x=LMARGIN`` avoids that entirely.
    """
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(0, h, text, align=align, new_x="LMARGIN", new_y="NEXT")


def _latin1(text) -> str:
    """Coerce arbitrary text to a latin-1-safe string for the PDF core fonts."""
    s = "" if text is None else str(text)
    for bad, good in _REPLACEMENTS.items():
        s = s.replace(bad, good)
    return s.encode("latin-1", "replace").decode("latin-1")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def excel_from_table(
    *, title: str, headers: list[str], rows: list[list],
    meta: dict[str, str] | None = None, sheet_name: str = "Reporte",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    r = 2
    for label, value in (meta or {}).items():
        ws.cell(row=r, column=1, value=f"{label}:").font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
        r += 1
    r += 1

    header_row = r
    for col, head in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=head)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for data_row in rows:
        r += 1
        for col, value in enumerate(data_row, start=1):
            ws.cell(row=r, column=col, value=value)

    # Auto width (bounded).
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        longest = len(str(headers[col - 1]))
        for data_row in rows:
            if col - 1 < len(data_row):
                longest = max(longest, len(str(data_row[col - 1])))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 48)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF (tables)
# ---------------------------------------------------------------------------
class _ReportPDF(FPDF):
    doc_title = ""

    def header(self) -> None:
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(120)
        half = (self.w - self.l_margin - self.r_margin) / 2
        self.cell(half, 6, _latin1(settings.institution_name), align="L")
        self.cell(half, 6, _latin1(f"Generado: {datetime.now():%d/%m/%Y %H:%M}"), align="R")
        self.ln(8)
        self.set_text_color(0)

    def footer(self) -> None:
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120)
        self.cell(0, 8, _latin1(f"Página {self.page_no()}/{{nb}}"), align="C")
        self.set_text_color(0)


def pdf_from_table(
    *, title: str, headers: list[str], rows: list[list],
    meta: dict[str, str] | None = None, landscape: bool = True,
) -> bytes:
    pdf = _ReportPDF(orientation="L" if landscape else "P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(31, 59, 87)
    _mc(pdf, 8, _latin1(title))
    pdf.set_text_color(0)
    pdf.set_font("Helvetica", "", 9)
    for label, value in (meta or {}).items():
        _mc(pdf, 5, _latin1(f"{label}: {value}"))
    pdf.ln(2)

    epw = pdf.w - pdf.l_margin - pdf.r_margin
    ncol = max(len(headers), 1)
    col_w = epw / ncol

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(31, 59, 87)
    pdf.set_text_color(255)
    for head in headers:
        pdf.cell(col_w, 7, _latin1(str(head))[:40], border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0)
    pdf.set_font("Helvetica", "", 8)
    fill = False
    for row in rows:
        pdf.set_fill_color(240, 244, 248)
        for col in range(ncol):
            value = row[col] if col < len(row) else ""
            pdf.cell(col_w, 6, _latin1(str(value))[:45], border=1, fill=fill)
        pdf.ln()
        fill = not fill
    if not rows:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 8, _latin1("Sin datos para los filtros seleccionados."))
    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Formal single-document PDF (letter layout — resignation reference structure)
# ---------------------------------------------------------------------------
def document_to_pdf(doc, *, type_label: str, student_name: str | None = None,
                    sede_name: str | None = None) -> bytes:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    # Institution header.
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 6, _latin1(settings.institution_name.upper()), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, _latin1("ESCUELA PROFESIONAL DE MEDICINA"), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Place, date and document number.
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, _latin1(f"Ñaña, Lima, {datetime.now():%d de %B del %Y}"), align="R",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _latin1(f"DOCUMENTO N° {doc.code}"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Recipient (destination).
    pdf.set_font("Helvetica", "B", 10)
    _mc(pdf, 5, _latin1(doc.destination or "Destinatario institucional"))
    pdf.ln(1)

    # Subject.
    pdf.set_font("Helvetica", "B", 10)
    _mc(pdf, 5, _latin1(f"Asunto: {doc.subject or doc.title}"))
    pdf.ln(1)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(90)
    _mc(pdf, 5, _latin1(f"Tipo: {type_label}  ·  Estado: {doc.status}  ·  Prioridad: {doc.priority}"))
    pdf.set_text_color(0)
    pdf.ln(3)

    # Greeting.
    pdf.set_font("Helvetica", "", 10)
    _mc(pdf, 5, _latin1("De mi mayor consideración:"))
    pdf.ln(2)

    # Context.
    ctx = []
    if student_name:
        ctx.append(f"Interno(a): {student_name}")
    if sede_name:
        ctx.append(f"Sede: {sede_name}")
    if ctx:
        pdf.set_font("Helvetica", "", 10)
        _mc(pdf, 5, _latin1("  ·  ".join(ctx)))
        pdf.ln(2)

    # Body.
    pdf.set_font("Helvetica", "", 10)
    _mc(pdf, 6, _latin1(doc.body or doc.summary or ""))
    pdf.ln(6)

    # Formal closing + signatory placeholder.
    _mc(pdf, 5, _latin1("Sin otro particular, quedo de usted."))
    pdf.ln(2)
    _mc(pdf, 5, _latin1("Atentamente,"))
    pdf.ln(12)
    pdf.cell(0, 5, _latin1(doc.origin or "________________________"), align="C",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120)
    _mc(pdf, 4, _latin1(
        "Documento generado por UPeU Internado 360 para gestión interna. "
        "Requiere revisión y aprobación humana; no constituye envío automático."))
    return bytes(pdf.output())
