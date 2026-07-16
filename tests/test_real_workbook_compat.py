"""Compatibility tests for the official "BASE DE DATOS - NOTAS INTERNADO
MÉDICO" workbook structure (Batch 2F patch).

Uses a SYNTHETIC workbook that mirrors the real file's structure exactly
(merged category band on the row above the header, header row 2, data from
row 3, sheet names ``QX 2026`` / ``INT. CIRUGÍA`` / etc.) — no real student
data is used, per SECURITY_AND_PRIVACY_RULES.md.
"""

from __future__ import annotations

import io
import json

import openpyxl

from app.database import SessionLocal
from app.models.grades import GradeComponentDefinition, GradeScheme
from app.repositories.repositories import RepositoryBundle
from app.services import excel_reader
from app.services.auth_service import Identity
from app.services.grade_service import (
    GradeService,
    category_code_for_band,
    rotation_hint_for_sheet,
)
from app.services.import_service import ImportService

_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _uni() -> Identity:
    db = SessionLocal()
    u = RepositoryBundle(db).users.get_by_email("coordinator@internado360.demo")
    ident = Identity(user_id=u.id, email=u.email, full_name=u.full_name,
                     role_code=u.role_code, role_name=u.role.name if u.role else "")
    db.close()
    return ident


def _codes(n=3):
    db = SessionLocal()
    codes = [s.student_code for s in RepositoryBundle(db).students.search()[:n]]
    db.close()
    return codes


def _build_real_structure_workbook(sheet_name: str, student_rows: list[list],
                                   comp_headers: list[str], categories: list[str]):
    """Mirror the real file: row1 title (merged), row2 category band (merged
    per span), row3 real headers (#, Código, Nombre del alumno, <components>),
    data from row4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    n_cols = 3 + len(comp_headers)
    ws.append(["Título"] + [None] * (n_cols - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Row 2: category band, merged per contiguous same-category run.
    band_row = [None, None, None] + categories
    ws.append(band_row)
    col = 4
    i = 0
    while i < len(categories):
        j = i
        while j + 1 < len(categories) and categories[j + 1] == categories[i]:
            j += 1
        if j > i:
            ws.merge_cells(start_row=2, start_column=col + i, end_row=2, end_column=col + j)
        i = j + 1

    ws.append(["#", "Código", "Nombre del alumno"] + comp_headers)
    for row in student_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Header/merged-band detection --------------------------------------------
def test_header_row_detected_below_merged_category_band():
    raw = _build_real_structure_workbook(
        "INT. CIRUGÍA", [[1, "S001", "Interno Uno", 15, 16, 0]],
        ["Lista de Cotejo - Ética", "Valoración del Proceso", "Exámenes Escritos"],
        ["Actitudinal", "Desempeño", "Conocimiento"])
    preview = excel_reader.read_sheet(raw, "INT. CIRUGÍA")
    assert preview.header_row_index == 2
    assert preview.headers[1] == "Código" and preview.headers[2] == "Nombre del alumno"


def test_category_band_forward_filled_across_merge():
    raw = _build_real_structure_workbook(
        "INT. MEDICINA", [[1, "S001", "Interno Uno", 15, 16, 14, 0, 17]],
        ["Lista de Cotejo", "Valoración A", "Valoración B", "Producto", "Examen"],
        ["Actitudinal", "Desempeño", "Desempeño", "Desempeño", "Conocimiento"])
    preview = excel_reader.read_sheet(raw, "INT. MEDICINA")
    # First 3 columns (#, Código, Nombre) have no band; components inherit
    # their merged category label.
    assert preview.category_headers[0:3] == ["", "", ""]
    assert preview.category_headers[3] == "Actitudinal"
    assert preview.category_headers[4] == "Desempeño"
    assert preview.category_headers[5] == "Desempeño"
    assert preview.category_headers[6] == "Desempeño"
    assert preview.category_headers[7] == "Conocimiento"


def test_category_code_mapping():
    assert category_code_for_band("Actitudinal") == "actitudinal"
    assert category_code_for_band("Desempeño") == "desempeno"
    assert category_code_for_band("Conocimiento") == "conocimiento"
    assert category_code_for_band("Algo Desconocido") == "otro"


def test_sheet_rotation_hints_match_real_sheet_names():
    assert rotation_hint_for_sheet("INT. CIRUGÍA") == "CIR"
    assert rotation_hint_for_sheet("INT. MEDICINA") == "MED"
    assert rotation_hint_for_sheet("INT. PEDIATRÍA") == "PED"
    assert rotation_hint_for_sheet("INT. GO") == "GO"
    assert rotation_hint_for_sheet("QX 2026") == ""
    assert rotation_hint_for_sheet("REV. MED. QUIR III") == ""


# --- Duplicate sheet names across a workbook ---------------------------------
def test_all_seven_expected_sheet_names_readable():
    """The real workbook ships exactly these 7 sheets; verify each is a valid,
    independently readable target (no duplicate-name collisions)."""
    names = ["QX 2026", "INT. CIRUGÍA", "INT. MEDICINA", "REV. MED. QUIR III",
            "INT. PEDIATRÍA", "INT. GO", "REV. MED. QUIR IV"]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for n in names:
        ws = wb.create_sheet(n)
        ws.append(["#", "Código", "Nombre del alumno", "Nota"])
        ws.append([1, "S001", "Interno Uno", 15])
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    assert excel_reader.list_sheets(raw) == names
    for n in names:
        p = excel_reader.read_sheet(raw, n)
        assert p.headers[1] == "Código"


# --- Student-key auto-mapping on the real header text ------------------------
def test_student_key_and_name_automap_on_real_headers():
    from app.services.import_profiles import get_profile
    gp = get_profile("grade_components")
    headers = ["#", "Código", "Nombre del alumno", "Lista de Cotejo - Ética"]
    mapping = gp.auto_map(headers)
    assert mapping["student_key"] == "Código"
    assert mapping["student_name"] == "Nombre del alumno"


# --- End-to-end grade import using the real structure -------------------------
_scheme_counter = [0]


def _scheme_from_headers(comp_headers, categories):
    _scheme_counter[0] += 1
    db = SessionLocal()
    r = RepositoryBundle(db)
    scheme = GradeScheme(code=f"GS-REAL-{_scheme_counter[0]}", name="Esquema real",
                         status="active", weights_confirmed=False)
    r.grade_schemes.add(scheme)
    db.flush()
    comp_ids = []
    for i, (h, cat) in enumerate(zip(comp_headers, categories)):
        c = GradeComponentDefinition(scheme_id=scheme.id, name=h,
                                     category=category_code_for_band(cat), display_order=i)
        r.grade_components.add(c)
        comp_ids.append(c)
    db.commit()
    sid, cids = scheme.id, [c.id for c in comp_ids]
    db.close()
    return sid, cids


def test_real_structure_import_blank_and_zero_preserved():
    codes = _codes(2)
    comp_headers = ["Lista de Cotejo", "Exámenes Escritos"]
    categories = ["Actitudinal", "Conocimiento"]
    sid, cids = _scheme_from_headers(comp_headers, categories)
    raw = _build_real_structure_workbook(
        "INT. CIRUGÍA",
        [[1, codes[0], "Interno Uno", 0, None],       # zero + blank
         [2, codes[1], "Interno Dos", 18, 15]],
        comp_headers, categories)

    db = SessionLocal()
    svc = ImportService(db, _uni())
    batch = svc.create_batch("grade_components", "notas.xlsx", _CT, raw=raw, scheme_id=sid)
    svc.set_sheet(batch.id, "INT. CIRUGÍA")
    mapping = {"student_key": "Código", "student_name": "Nombre del alumno",
               f"comp_{cids[0]}": comp_headers[0], f"comp_{cids[1]}": comp_headers[1],
               "_scheme_id": str(sid)}
    svc.set_mapping(batch.id, mapping, "valid_only")
    svc.validate_batch(batch.id)
    svc.confirm_batch(batch.id)
    r = RepositoryBundle(db)
    st0 = r.students.get_by_code(codes[0])
    sgc_zero = r.student_grades.get_one(st0.id, sid, cids[0])
    sgc_blank = r.student_grades.get_one(st0.id, sid, cids[1])
    db.close()
    assert sgc_zero.score == 0.0
    assert sgc_blank.score is None
    assert sgc_zero.source_sheet == "INT. CIRUGÍA"
    assert sgc_zero.source_col == comp_headers[0]


def test_real_structure_missing_student_flagged():
    comp_headers = ["Lista de Cotejo"]
    categories = ["Actitudinal"]
    sid, cids = _scheme_from_headers(comp_headers, categories)
    raw = _build_real_structure_workbook(
        "INT. GO", [[1, "NOEXISTE999", "Fantasma", 10]], comp_headers, categories)
    db = SessionLocal()
    svc = ImportService(db, _uni())
    batch = svc.create_batch("grade_components", "notas.xlsx", _CT, raw=raw, scheme_id=sid)
    svc.set_sheet(batch.id, "INT. GO")
    mapping = {"student_key": "Código", "student_name": "Nombre del alumno",
              f"comp_{cids[0]}": comp_headers[0], "_scheme_id": str(sid)}
    svc.set_mapping(batch.id, mapping, "valid_only")
    svc.validate_batch(batch.id)
    rows = svc.repos.import_rows.for_batch(batch.id)
    messages = json.loads(rows[0].messages_json or "[]")
    db.close()
    assert rows[0].status == "error"
    assert any("NOEXISTE999" in m["message"] for m in messages)


# --- Cross-sheet inconsistency detection --------------------------------------
def test_cross_sheet_missing_student_detected():
    codes = _codes(3)
    comp_headers, categories = ["Nota"], ["Conocimiento"]
    sid, cids = _scheme_from_headers(comp_headers, categories)
    raw_a = _build_real_structure_workbook(
        "QX 2026", [[1, codes[0], "A", 15], [2, codes[1], "B", 16]], comp_headers, categories)
    raw_b = _build_real_structure_workbook(
        "INT. CIRUGÍA", [[1, codes[1], "B", 14], [2, codes[2], "C", 12]], comp_headers, categories)

    db = SessionLocal()
    svc = ImportService(db, _uni())
    mapping = lambda: {"student_key": "Código", "student_name": "Nombre del alumno",
                       f"comp_{cids[0]}": comp_headers[0], "_scheme_id": str(sid)}
    bA = svc.create_batch("grade_components", "a.xlsx", _CT, raw=raw_a, scheme_id=sid)
    svc.set_sheet(bA.id, "QX 2026")
    svc.set_mapping(bA.id, mapping(), "valid_only")
    svc.validate_batch(bA.id); svc.confirm_batch(bA.id)

    bB = svc.create_batch("grade_components", "b.xlsx", _CT, raw=raw_b, scheme_id=sid)
    svc.set_sheet(bB.id, "INT. CIRUGÍA")
    svc.set_mapping(bB.id, mapping(), "valid_only")
    svc.validate_batch(bB.id); svc.confirm_batch(bB.id)

    gsvc = GradeService(db, _uni())
    report = gsvc.cross_sheet_report([bA.id, bB.id])
    missing_keys = {m["student_key"] for m in report["missing"]}
    db.close()
    assert codes[0] in missing_keys  # present only in QX 2026
    assert codes[2] in missing_keys  # present only in INT. CIRUGÍA
    assert codes[1] not in missing_keys  # present in both


def test_cross_sheet_name_mismatch_detected():
    codes = _codes(1)
    comp_headers, categories = ["Nota"], ["Conocimiento"]
    sid, cids = _scheme_from_headers(comp_headers, categories)
    raw_a = _build_real_structure_workbook(
        "QX 2026", [[1, codes[0], "Nombre Correcto", 15]], comp_headers, categories)
    raw_b = _build_real_structure_workbook(
        "INT. MEDICINA", [[1, codes[0], "NOMBRE DIFERENTE", 14]], comp_headers, categories)

    db = SessionLocal()
    svc = ImportService(db, _uni())
    mapping = lambda: {"student_key": "Código", "student_name": "Nombre del alumno",
                       f"comp_{cids[0]}": comp_headers[0], "_scheme_id": str(sid)}
    bA = svc.create_batch("grade_components", "a.xlsx", _CT, raw=raw_a, scheme_id=sid)
    svc.set_sheet(bA.id, "QX 2026")
    svc.set_mapping(bA.id, mapping(), "valid_only")
    svc.validate_batch(bA.id); svc.confirm_batch(bA.id)

    bB = svc.create_batch("grade_components", "b.xlsx", _CT, raw=raw_b, scheme_id=sid)
    svc.set_sheet(bB.id, "INT. MEDICINA")
    svc.set_mapping(bB.id, mapping(), "valid_only")
    svc.validate_batch(bB.id); svc.confirm_batch(bB.id)

    gsvc = GradeService(db, _uni())
    report = gsvc.cross_sheet_report([bA.id, bB.id])
    db.close()
    assert any(m["student_key"] == codes[0] for m in report["mismatches"])


def test_cross_sheet_no_false_positive_when_rosters_match():
    codes = _codes(2)
    comp_headers, categories = ["Nota"], ["Conocimiento"]
    sid, cids = _scheme_from_headers(comp_headers, categories)
    raw_a = _build_real_structure_workbook(
        "QX 2026", [[1, codes[0], "A", 15], [2, codes[1], "B", 16]], comp_headers, categories)
    raw_b = _build_real_structure_workbook(
        "INT. GO", [[1, codes[0], "A", 14], [2, codes[1], "B", 12]], comp_headers, categories)

    db = SessionLocal()
    svc = ImportService(db, _uni())
    mapping = lambda: {"student_key": "Código", "student_name": "Nombre del alumno",
                       f"comp_{cids[0]}": comp_headers[0], "_scheme_id": str(sid)}
    bA = svc.create_batch("grade_components", "a.xlsx", _CT, raw=raw_a, scheme_id=sid)
    svc.set_sheet(bA.id, "QX 2026")
    svc.set_mapping(bA.id, mapping(), "valid_only")
    svc.validate_batch(bA.id); svc.confirm_batch(bA.id)

    bB = svc.create_batch("grade_components", "b.xlsx", _CT, raw=raw_b, scheme_id=sid)
    svc.set_sheet(bB.id, "INT. GO")
    svc.set_mapping(bB.id, mapping(), "valid_only")
    svc.validate_batch(bB.id); svc.confirm_batch(bB.id)

    gsvc = GradeService(db, _uni())
    report = gsvc.cross_sheet_report([bA.id, bB.id])
    db.close()
    assert report["missing"] == []
    assert report["mismatches"] == []


# --- Route registration order (static path before /grades/{scheme_id}) ------
def test_cross_sheet_check_route_not_shadowed_by_scheme_detail(university_client):
    """/grades/cross-sheet-check must resolve to the dedicated view, not be
    swallowed by /grades/{scheme_id} (which would 422 on a non-integer id)."""
    r = university_client.get("/grades/cross-sheet-check")
    assert r.status_code == 200
    assert "Verificaci" in r.text  # "Verificación entre hojas"


# --- Approved-grade protection preserved with the real structure -------------
def test_approved_grade_not_silently_overwritten_real_structure():
    codes = _codes(1)
    comp_headers, categories = ["Nota"], ["Conocimiento"]
    sid, cids = _scheme_from_headers(comp_headers, categories)
    raw1 = _build_real_structure_workbook(
        "QX 2026", [[1, codes[0], "A", 15]], comp_headers, categories)
    db = SessionLocal()
    svc = ImportService(db, _uni())
    mapping = lambda: {"student_key": "Código", "student_name": "Nombre del alumno",
                       f"comp_{cids[0]}": comp_headers[0], "_scheme_id": str(sid)}
    b1 = svc.create_batch("grade_components", "a.xlsx", _CT, raw=raw1, scheme_id=sid)
    svc.set_sheet(b1.id, "QX 2026")
    svc.set_mapping(b1.id, mapping(), "valid_only")
    svc.validate_batch(b1.id); svc.confirm_batch(b1.id)

    r = RepositoryBundle(db)
    st = r.students.get_by_code(codes[0])
    sgc = r.student_grades.get_one(st.id, sid, cids[0])
    GradeService(db, _uni()).approve_component(sgc.id)

    raw2 = _build_real_structure_workbook(
        "QX 2026", [[1, codes[0], "A", 9]], comp_headers, categories)
    b2 = svc.create_batch("grade_components", "a2.xlsx", _CT, raw=raw2, scheme_id=sid)
    svc.set_sheet(b2.id, "QX 2026")
    svc.set_mapping(b2.id, mapping(), "valid_only")
    svc.validate_batch(b2.id)
    rows = svc.repos.import_rows.for_batch(b2.id)
    messages = json.loads(rows[0].messages_json or "[]")
    assert any("aprobada" in m["message"] for m in messages)
    db.close()
